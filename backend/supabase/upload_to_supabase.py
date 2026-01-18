import os
import glob
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl import load_workbook

# 1. Load env
load_dotenv()
URL = os.environ.get("SUPABASE_URL")
KEY = os.environ.get("SUPABASE_KEY")

if not URL or not KEY:
    print("❌ Error: Check your .env file for SUPABASE_URL and SUPABASE_KEY")
    exit()

supabase: Client = create_client(URL, KEY)

# Map Excel Headers -> Supabase Columns
HEADER_MAP = {
    "Grant Name": "grant_name",
    "URL": "url",
    "Description": "description",
    "Issue Area": "issue_area",
    "Scope": "scope",
    "KPIs": "kpis",
    "Funding": "funding",
    "Due Date": "due_date"
}

def get_latest_excel():
    search_paths = ["backend/scraper/output/Gemini_Grant_Report_20260116_010612.xlsx",]
    found_files = []
    for path in search_paths:
        found_files.extend(glob.glob(path))
    
    if not found_files:
        return None
    return max(found_files, key=os.path.getmtime)

def main():
    file_path = get_latest_excel()
    if not file_path:
        print("❌ No Excel file found.")
        return

    print(f"Reading: {file_path}")
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # 1. Get Headers from the first row
    headers = [cell.value for cell in ws[1]]
    
    records = []
    
    # 2. Loop through the rest of the rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        # Zip creates pairs of (Header, Value)
        for header, value in zip(headers, row):
            # Only process if header is in our map
            if header in HEADER_MAP:
                db_col = HEADER_MAP[header]
                # Convert Excel empty cells (None) stays None, which Supabase likes
                row_data[db_col] = value
        
        # Only add if we actually have a grant name (skip empty rows)
        if row_data.get("grant_name"):
            records.append(row_data)

    print(f"Uploading {len(records)} grants...")
    
    try:
        supabase.table("grants").upsert(
            records, 
            on_conflict="grant_name"
        ).execute()
        print("✅ Success!")
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()